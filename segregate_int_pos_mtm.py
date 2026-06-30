"""
Intraday / Positional segregation report.

Logic
-----
1. Read every account from the Updated Compiled User MTM file.
2. An account is POSITIONAL if its User ID appears in EITHER of the two
   "Combined max loss Calculation" files (23-06 / 20-06); otherwise INTRADAY.
3. For a positional account the report's Realized P&L is:
       compiled Realized P&L
     + (Realized PNL + Net Settlement Value) from the 23-06 file (if present)
     + (Realized PNL + Net Settlement Value) from the 20-06 file (if present)
   i.e. summed across BOTH files when the account is in both.
   Intraday accounts keep the plain compiled Realized P&L.
4. Output is a brand-new workbook, grouped per ALGO, and inside each algo split
   into an Int block and a Pos+Int block, each with its own Sub-Total,
   followed by a single overall Grand Total.
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ====================== PATHS ======================
BASE = r"D:\Gk\work\0dte testinging"
compiled_path = os.path.join(BASE, "Updated_Compiled_User_MTM_23-06-2026.xlsx")


# ====================== ASK FOR INPUT FILES ======================
def ask_file(label, required):
    while True:
        raw = input(f"  >> {label}\n     Path: ").strip().strip('"').strip("'").strip()
        if not raw:
            if required:
                print("     [!] This file is required - please provide a path.\n")
                continue
            print("     [i] Skipped (no file given).\n")
            return None
        if not os.path.exists(raw):
            print(f"     [!] File not found: {raw}\n")
            if required:
                continue
            again = input("     Type a path to retry, or press Enter to skip: ").strip().strip('"').strip("'")
            if not again:
                print("     [i] Skipped.\n")
                return None
            if os.path.exists(again):
                return again
            print("     [!] Still not found - skipping.\n")
            return None
        return raw


print("=" * 60)
print("  COMBINED MAX LOSS - FILE UPLOAD")
print("  (Both files optional, but at least the 1DTE file is needed.)")
print("=" * 60)
path_4dte = ask_file("Upload 4DTE Combined max loss file (press Enter to skip):", required=False)
path_1dte = ask_file("Upload 1DTE Combined max loss file:", required=True)


# ====================== LOAD DATA ======================
def load_combined(path):
    if not path:
        return {}
    df = pd.read_excel(path, sheet_name="Sheet1")
    df["User ID"] = df["User ID"].astype(str).str.strip()
    df["Realized PNL"] = pd.to_numeric(df["Realized PNL"], errors="coerce").fillna(0)
    df["Net Settlement Value"] = pd.to_numeric(df["Net Settlement Value"], errors="coerce").fillna(0)
    df["User Type"] = df["User Type"].astype(str).str.strip()
    df = df.drop_duplicates(subset=["User ID"], keep="last")
    out = {}
    for _, r in df.iterrows():
        out[r["User ID"]] = {
            "addon": float(r["Realized PNL"] + r["Net Settlement Value"]),
            "type": r["User Type"],
        }
    return out


print("\nReading source files...")
comp = pd.read_excel(compiled_path, sheet_name="Sheet1")
comp["UserID"] = comp["UserID"].astype(str).str.strip()

for col in ["Realized P&L", "ALLOCATION", "MAX LOSS", "Unrealized P&L"]:
    if col in comp.columns:
        comp[col] = pd.to_numeric(comp[col], errors="coerce").fillna(0)
    else:
        comp[col] = 0

four = load_combined(path_4dte)
one  = load_combined(path_1dte)
positional_ids = set(four) | set(one)


def user_type(uid):
    if uid in one:
        return one[uid]["type"]
    if uid in four:
        return four[uid]["type"]
    return ""


def applied_addon(uid):
    a1 = one.get(uid, {}).get("addon", 0.0)
    a4 = four.get(uid, {}).get("addon", 0.0)
    if str(user_type(uid)).strip().lower() == "noren":
        return a1
    return a1 + a4


comp["Type"]         = comp["UserID"].apply(lambda u: "Positional" if u in positional_ids else "Intraday")
comp["User Type"]    = comp["UserID"].apply(lambda u: user_type(u) if u in positional_ids else "")
comp["Addon 4DTE"]   = comp["UserID"].apply(lambda u: four.get(u, {}).get("addon", 0.0))
comp["Addon 1DTE"]   = comp["UserID"].apply(lambda u: one.get(u, {}).get("addon", 0.0))
comp["Addon Applied"]= comp.apply(
    lambda r: applied_addon(r["UserID"]) if r["Type"] == "Positional" else 0.0, axis=1
)
comp["AdjRealized"]  = comp["Realized P&L"] + comp["Addon Applied"]
# Per-user return (Realized / Allocation) — basis for the P5/P95 percentile columns.
comp["UserReturn"]   = comp.apply(
    lambda r: (r["AdjRealized"] / r["ALLOCATION"]) if r["ALLOCATION"] else 0.0, axis=1
)

REPORT_DATE = "23-06-2026"
if "Date" in comp.columns and comp["Date"].notna().any():
    REPORT_DATE = str(comp["Date"].dropna().iloc[0]).split(" ")[0]
output_path = os.path.join(BASE, f"Segregated_Int_Pos_MTM_{REPORT_DATE}.xlsx")

n_pos   = (comp["Type"] == "Positional").sum()
n_noren = (comp["User Type"].str.lower() == "noren").sum()
print(f"  Files loaded -> 4DTE: {'yes' if four else 'NO'} | 1DTE: {'yes' if one else 'NO'}")
print(f"  Accounts: {len(comp)}  |  Positional: {n_pos}  |  Intraday: {len(comp) - n_pos}")
print(f"  Positional breakdown -> Noren: {n_noren} | Non-Noren: {n_pos - n_noren}")


# ====================== AGGREGATION ======================
def pct_returns(df_block: pd.DataFrame):
    """5th / 95th percentile of the per-user returns inside a group.

    P5/P95 stand in for min/max while controlling for data aberrations
    (intraday pauses/stops, under/over-funded accounts, suboptimal allocations).
    Users with zero allocation have an undefined return and are excluded.
    Returns (p5, p95).
    """
    rr = df_block.loc[df_block["ALLOCATION"] > 0, "UserReturn"].dropna()
    if rr.empty:
        return (0.0, 0.0)
    return (float(rr.quantile(0.05)), float(rr.quantile(0.95)))


def aggregate(sub: pd.DataFrame):
    rows = []
    for server, g in sub.groupby("SERVER"):
        realized = g["AdjRealized"].sum()
        unreal   = g["Unrealized P&L"].sum()
        alloc    = g["ALLOCATION"].sum()
        p5, p95  = pct_returns(g)
        rows.append({
            "ALGO"      : g["ALGO"].iloc[0],
            "SERVER"    : server,
            "Users"     : len(g),
            "SLHit"     : int((g["SL HIT/NOT"] == 1).sum()),
            "MaxLoss"   : g["MAX LOSS"].sum(),
            "Allocation": alloc,
            "Realized"  : realized,
            "Unrealized": unreal,
            "MTM"       : realized + unreal,
            "Return"    : (realized / alloc) if alloc else 0.0,
            "P95"       : p95,
            "P5"        : p5,
        })
    rows.sort(key=lambda x: str(x["SERVER"]))
    return rows


# ====================== STYLES ======================
HEADERS = ["ALGO", "SERVER", "No. of Users", "No. of SL Hit Users", "MAX LOSS",
           "ALLOCATION", "Realized P&L", "Unrealized P&L", "MTM", "Return %",
           "95%", "5%"]
KEYS    = ["ALGO", "SERVER", "Users", "SLHit", "MaxLoss",
           "Allocation", "Realized", "Unrealized", "MTM", "Return",
           "P95", "P5"]
NCOLS   = 1 + len(HEADERS)   # col A = type label, cols B-M = data

# Column letters (A=1, B=2, ..., M=13)
# B=ALGO  C=SERVER  D=Users  E=SLHit  F=MaxLoss
# G=Allocation  H=Realized  I=Unrealized  J=MTM  K=Return  L=P95  M=P5
KEY_COL   = {"Users": "D", "SLHit": "E", "MaxLoss": "F",
             "Allocation": "G", "Realized": "H", "Unrealized": "I", "MTM": "J"}
MONEY_KEYS = {"MaxLoss", "Allocation", "Realized", "Unrealized", "MTM"}
PCT_KEYS   = {"P95", "P5"}   # 5th/95th percentile of per-user returns (written as values)

# --- Border sides ---
_thin = Side(style="thin",   color="C0C0C0")
_med  = Side(style="medium", color="595959")

def _bdr(l=None, r=None, t=None, b=None):
    return Border(left=l or _thin, right=r or _thin, top=t or _thin, bottom=b or _thin)

inner_bdr    = _bdr()
sub_bdr      = _bdr(t=_med, b=_med)     # medium top + bottom — sub-total reads as separator bar
banner_bdr   = _bdr(t=_med)             # top edge of each algo banner
algo_tot_bdr = _bdr(t=_med, b=_med)    # algo total: medium top + bottom
grand_bdr    = Border(left=_med, right=_med, top=_med, bottom=_med)

# --- Alignments ---
center      = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
vertical_a  = Alignment(horizontal="center", vertical="center", text_rotation=90)

# --- Fills ---
# Only title + header are dark. Every algo cycles through its own light palette pair.
title_fill  = PatternFill("solid", fgColor="1E293B")  # Deep Slate Blue/Gray  (title bar)
header_fill = PatternFill("solid", fgColor="64748B")  # Muted Cool Gray (column headers)
grand_fill  = PatternFill("solid", fgColor="D1C9E1")  # light lavender (grand total)

# Per-algo rotating palette — (data_hex, accent_hex)
#   data_hex   : ~10-12% saturation, used for ALL data rows in the algo
#   accent_hex : slightly more saturated, used for the banner, col-A label, algo-total row
#   sub-total  : derived = accent lightened 60% toward white, so it sits between the
#                data rows (lightest) and the accent, giving it its own distinct shade
ALGO_PALETTE = [
    ("FCE8E6", "E8A898"),  # blush pink
    ("E8F2E4", "98C490"),  # sage green
    ("EAE8F8", "9898D8"),  # periwinkle
    ("E4F0F8", "80BCDC"),  # sky blue
    ("F8F4E4", "C8B868"),  # warm cream / amber
    ("E8F6F2", "78BCA8"),  # seafoam / mint
    ("F8E8EE", "D898B0"),  # rose pink
    ("EEE8F8", "B090D0"),  # soft lilac
    ("E4F6F8", "68B0B8"),  # pale teal
]

def _lighten(hex_color, factor):
    """Lighten a hex colour by blending it `factor` of the way toward white."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = round(r + (255 - r) * factor)
    g = round(g + (255 - g) * factor)
    b = round(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"

def _algo_fills(i):
    data_hex, accent_hex = ALGO_PALETTE[i % len(ALGO_PALETTE)]
    sub_hex = _lighten(accent_hex, 0.60)   # sub-total: 60% lighter version of the accent
    return (PatternFill("solid", fgColor=data_hex),
            PatternFill("solid", fgColor=accent_hex),
            PatternFill("solid", fgColor=sub_hex))

white_bold = Font(bold=True, color="FFFFFF")
bold       = Font(bold=True)


# ====================== WORKBOOK ======================
wb = Workbook()
ws = wb.active
ws.title = "Segregation"

# Title row
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
tc = ws.cell(1, 1, f"Algo & Server wise Realized P&L  —  Int / Pos+Int  ({REPORT_DATE})")
tc.font      = Font(bold=True, italic=True, color="FFFFFF", size=13)
tc.fill      = title_fill
tc.alignment = center

# Header row
HEADER_ROW = 2
ws.cell(HEADER_ROW, 1, "Type")
for j, h in enumerate(HEADERS):
    ws.cell(HEADER_ROW, j + 2, h)
for c in range(1, NCOLS + 1):
    cell           = ws.cell(HEADER_ROW, c)
    cell.fill      = header_fill
    cell.font      = white_bold
    cell.alignment = center_wrap
    cell.border    = inner_bdr


# ====================== ROW WRITERS ======================
def _base(cell, fill, bdr, is_bold=False):
    cell.fill      = fill
    cell.border    = bdr
    cell.alignment = center
    if is_bold:
        cell.font = bold


def _algo_val(v):
    """Display algo number as integer when it is a whole float."""
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def write_data_row(row, rowdict, fill, is_bold=False, bdr=None):
    b = bdr or inner_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "Return":
            cell.value         = f"=IF(G{row}=0,0,H{row}/G{row})"
            cell.number_format = "0.00"
        elif key == "ALGO":
            cell.value = _algo_val(rowdict[key])
        elif key in MONEY_KEYS:
            cell.value         = round(float(rowdict[key]))
            cell.number_format = "#,##0"
        elif key in PCT_KEYS:
            cell.value         = round(float(rowdict[key]), 4)
            cell.number_format = "0.00"
        else:
            cell.value = rowdict[key]


def write_subtotal_row(row, data_start, data_end, fill, n_servers, pct=None, is_bold=True, bdr=None):
    """Sub-total row: SUM formulas for all aggregated columns.
    pct: {"P95":.., "P5":..} percentiles over all users in the section."""
    b = bdr or sub_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "ALGO":
            cell.value = "Sub-Total"
        elif key == "SERVER":
            cell.value = n_servers
        elif key in KEY_COL:
            cl             = KEY_COL[key]
            cell.value     = f"=SUM({cl}{data_start}:{cl}{data_end})"
            if key in MONEY_KEYS:
                cell.number_format = "#,##0"
        elif key == "Return":
            cell.value         = f"=IF(G{row}=0,0,H{row}/G{row})"
            cell.number_format = "0.00"
        elif key in PCT_KEYS and pct is not None:
            cell.value         = round(float(pct[key]), 4)
            cell.number_format = "0.00"


def write_total_row(row, label, n_servers, ref_rows, fill, pct=None, is_bold=True, bdr=None):
    """Algo-total or Grand-total: reference (sum) the given ref_rows.
    pct: {"P95":.., "P5":..} percentiles over all users in the algo / whole book."""
    b = bdr or algo_tot_bdr
    for j, key in enumerate(KEYS):
        col  = j + 2
        cell = ws.cell(row, col)
        _base(cell, fill, b, is_bold)
        if key == "ALGO":
            cell.value = label
        elif key == "SERVER":
            cell.value = n_servers
        elif key in KEY_COL:
            cl         = KEY_COL[key]
            formula    = "+".join(f"{cl}{rr}" for rr in ref_rows)
            cell.value = f"={formula}"
            if key in MONEY_KEYS:
                cell.number_format = "#,##0"
        elif key == "Return":
            cell.value         = f"=IF(G{row}=0,0,H{row}/G{row})"
            cell.number_format = "0.00"
        elif key in PCT_KEYS and pct is not None:
            cell.value         = round(float(pct[key]), 4)
            cell.number_format = "0.00"


def apply_outer_border(start_row, end_row, start_col=1, end_col=NCOLS):
    """Overlay a medium border on the outer perimeter of a cell range.
    Preserves all inner borders that are already set."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            b    = cell.border
            cell.border = Border(
                left   = _med if col == start_col else (b.left   or _thin),
                right  = _med if col == end_col   else (b.right  or _thin),
                top    = _med if row == start_row else (b.top    or _thin),
                bottom = _med if row == end_row   else (b.bottom or _thin),
            )


# ====================== BUILD SHEET ======================
r               = HEADER_ROW + 1
algo_total_rows = []    # row numbers of each algo-total row (for grand total formula)
total_n_servers = 0

algos = sorted(comp["ALGO"].dropna().unique(), key=lambda x: float(x))

for i, algo in enumerate(algos):
    algo_df        = comp[comp["ALGO"] == algo]
    algo_label     = f"Algo {int(algo) if float(algo).is_integer() else algo}"
    d_fill, a_fill, s_fill = _algo_fills(i)   # data + accent + sub-total fills for this algo

    # ---- Algo banner (accent colour, medium top border) ----
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOLS)
    bc           = ws.cell(r, 2, algo_label)
    bc.fill      = a_fill
    bc.font      = Font(bold=True, color="1F2D3D", size=11)
    bc.alignment = center
    for c in range(1, NCOLS + 1):
        cell        = ws.cell(r, c)
        cell.fill   = a_fill
        cell.border = banner_bdr
    ws.row_dimensions[r].height = 20
    r += 1

    section_sub_rows = []

    for ttype, display_label in [("Intraday", "Int"), ("Positional", "Pos+Int")]:
        sub = algo_df[algo_df["Type"] == ttype]
        if sub.empty:
            continue

        rows        = aggregate(sub)
        block_start = r

        for rowdict in rows:
            write_data_row(r, rowdict, d_fill)
            ws.row_dimensions[r].height = 18
            r += 1

        data_end = r - 1
        sub_row  = r
        # sub-total uses its own 60%-lighter accent shade so it reads distinct from data rows
        sub_p5, sub_p95 = pct_returns(sub)
        write_subtotal_row(r, block_start, data_end, s_fill, n_servers=len(rows),
                           pct={"P95": sub_p95, "P5": sub_p5})
        ws.row_dimensions[r].height = 18
        r += 1

        # Section label col A — merged, rotated, accent colour
        ws.merge_cells(start_row=block_start, start_column=1, end_row=sub_row, end_column=1)
        lc           = ws.cell(block_start, 1, display_label)
        lc.fill      = a_fill
        lc.font      = Font(bold=True, color="1F2D3D")
        lc.alignment = vertical_a
        for rr in range(block_start, sub_row + 1):
            ws.cell(rr, 1).fill   = a_fill
            ws.cell(rr, 1).border = inner_bdr

        # Outer medium border boxes the whole Int / Pos+Int section
        apply_outer_border(block_start, sub_row)

        section_sub_rows.append(sub_row)

    # Algo-level server count is DISTINCT across Int + Pos+Int: a server that runs
    # both types appears in each section but should be counted once for the algo.
    algo_n_servers = algo_df["SERVER"].nunique()

    # ---- Algo Total — col A:B merged & centered (no orphan empty cell) ----
    algo_p5, algo_p95 = pct_returns(algo_df)
    write_total_row(r, f"{algo_label} Total", algo_n_servers, section_sub_rows, a_fill,
                    pct={"P95": algo_p95, "P5": algo_p5})
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    mc           = ws.cell(r, 1)          # top-left cell owns the merged range
    mc.value     = f"{algo_label} Total"
    mc.fill      = a_fill
    mc.font      = bold
    mc.alignment = center
    mc.border    = algo_tot_bdr
    ws.row_dimensions[r].height = 20
    apply_outer_border(r, r)
    algo_total_rows.append(r)
    total_n_servers += algo_n_servers
    r += 1

    r += 1   # blank spacer row between algos

# ====================== GRAND TOTAL ======================
grand_p5, grand_p95 = pct_returns(comp)
write_total_row(r, "Grand Total", total_n_servers, algo_total_rows, grand_fill,
                pct={"P95": grand_p95, "P5": grand_p5}, bdr=grand_bdr)
gc           = ws.cell(r, 1, "Total")
gc.fill      = grand_fill
gc.font      = bold
gc.alignment = center
gc.border    = grand_bdr

# ---- Footnote explaining the P5/P95 (95% / 5%) columns ----
note_row = r + 2
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=NCOLS)
nc = ws.cell(
    note_row, 1,
    "*P5/P95 are the 5th and 95th percentile, and are used for representing min/max "
    "respectively. This is done to control for data aberrations due to exceptions like "
    "intraday account pause/stops (tech errors/funding issue/broker issue etc), returns "
    "due to underfunded or overfunded accounts, or statistically abnormal returns due to "
    "suboptimal allocations etc."
)
nc.font      = Font(italic=True, color="595959", size=9)
nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[note_row].height = 46

# ====================== FINISH FORMATTING ======================
ws.freeze_panes = f"A{HEADER_ROW + 1}"
ws.column_dimensions["A"].width = 5
widths = [7, 9, 12, 18, 13, 13, 14, 15, 13, 10, 9, 9]
for j, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(j + 2)].width = w
ws.row_dimensions[1].height = 22


# ====================== RAW PER-USER DATA SHEET ======================
print("Writing raw per-user data sheet...")
comp["MTM_calc"] = comp["AdjRealized"] + comp["Unrealized P&L"]
comp["SL Hit"]   = (comp["SL HIT/NOT"] == 1).astype(int)

raw_cols = [
    ("Type",                 "Type"),
    ("User Type",            "User Type"),
    ("ALGO",                 "ALGO"),
    ("SERVER",               "SERVER"),
    ("UserID",               "UserID"),
    ("Alias",                "Alias"),
    ("SL Hit",               "SL Hit"),
    ("MAX LOSS",             "MAX LOSS"),
    ("ALLOCATION",           "ALLOCATION"),
    ("Compiled Realized P&L","Realized P&L"),
    ("Addon 4DTE",           "Addon 4DTE"),
    ("Addon 1DTE",           "Addon 1DTE"),
    ("Addon Applied",        "Addon Applied"),
    ("Realized P&L (Final)", "AdjRealized"),
    ("Unrealized P&L",       "Unrealized P&L"),
    ("MTM",                  "MTM_calc"),
]
raw = comp[[src for _, src in raw_cols]].copy()
raw.columns = [hdr for hdr, _ in raw_cols]
raw = raw.sort_values(["ALGO", "Type", "SERVER", "UserID"]).reset_index(drop=True)

ws2 = wb.create_sheet("Raw_Data_Per_User")
for j, hdr in enumerate(raw.columns, start=1):
    cell           = ws2.cell(1, j, hdr)
    cell.fill      = header_fill
    cell.font      = white_bold
    cell.alignment = center_wrap
    cell.border    = inner_bdr

money_cols = {"MAX LOSS", "ALLOCATION", "Compiled Realized P&L",
              "Addon 4DTE", "Addon 1DTE", "Addon Applied",
              "Realized P&L (Final)", "Unrealized P&L", "MTM"}
for i, rec in enumerate(raw.itertuples(index=False), start=2):
    for j, (hdr, val) in enumerate(zip(raw.columns, rec), start=1):
        cell        = ws2.cell(i, j, val)
        cell.border = inner_bdr
        if hdr in money_cols and isinstance(val, (int, float)):
            cell.value         = float(val)
            cell.number_format = "#,##0"
        if hdr in ("Type", "ALGO", "SERVER", "SL Hit"):
            cell.alignment = center

ws2.freeze_panes = "A2"
raw_widths = [11, 7, 9, 12, 22, 7, 12, 12, 18, 18, 18, 15, 13]
for j, w in enumerate(raw_widths, start=1):
    ws2.column_dimensions[get_column_letter(j)].width = w

wb.save(output_path)

py_realized = comp["AdjRealized"].sum()
py_mtm      = (comp["AdjRealized"] + comp["Unrealized P&L"]).sum()
print(f"\nDone. Report saved at:\n  {output_path}")
print(f"  Grand Total -> Users: {len(comp)}  Realized P&L: {round(py_realized):,}  MTM: {round(py_mtm):,}")
