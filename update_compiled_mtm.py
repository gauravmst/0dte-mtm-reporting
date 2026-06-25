import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# use this code where compile user mtm jo daily compile se aata h wo wali file daalni h fr saved mtm wali file algo ui se jo file aati h wo h 
# or output path m jha save krna h wo thik h isse kya hoga ki jo saved mtm ka realized p&l or max loss 
# wo apne sheet m lg jaega sath m sheet2 bhi wo update kr dega thik 

# ====================== PATHS ======================
compiled_path = r"G:\My Drive\mtmformat\Compiled_User_MTM_23-06-2026.xlsx"
saved_path = r"G:\My Drive\mtmformat\saved_mtm_2026-06-23.xlsx"
output_path = r"G:\My Drive\mtmformat\Updated_Compiled_User_MTM_23-06-2026.xlsx"

# ====================== SHEET 1 UPDATE ======================
print("🔄 Updating Sheet1...")

# Load original workbook to preserve all sheets
wb = load_workbook(compiled_path)
ws1 = wb['Sheet1']

# Read into pandas for easy updating
compiled_df = pd.read_excel(compiled_path, sheet_name='Sheet1')
saved_df = pd.read_excel(saved_path, sheet_name='Sheet1')

# Skip blank early rows
saved_df = saved_df[saved_df['MTM'].notna()].copy()

# CC → XLDH Mapping
cc_to_xldh = {
    'CC04': 'XLDH142', 'CC05': 'XLDH158', 'CC09': 'XLDH159',
    'CC10': 'XLDH162', 'CC03': 'XLDH161', 'CC08': 'XLDH168',
}

user_map = {}
for cc, xldh in cc_to_xldh.items():
    user_map[cc] = xldh
    user_map[xldh] = xldh

saved_df = saved_df.drop_duplicates(subset=['user_id'], keep='last')

# Build mapping
mtm_map = {}
for _, row in saved_df.iterrows():
    uid = str(row['user_id']).strip()
    mapped_uid = user_map.get(uid, uid)
    
    realized = row.get('MTM', 0)
    max_loss = row.get('max_loss', 0)
    
    if pd.notna(max_loss) and max_loss < 0:
        max_loss = abs(max_loss)
    
    mtm_map[mapped_uid] = {'realized': realized, 'max_loss': max_loss}

# Apply updates
updated_count = 0
not_found = []
for idx, row in compiled_df.iterrows():
    user_id = str(row.get('UserID', '')).strip()
    if user_id and user_id in mtm_map:
        compiled_df.at[idx, 'Realized P&L'] = mtm_map[user_id]['realized']
        if 'MAX LOSS' in compiled_df.columns:
            compiled_df.at[idx, 'MAX LOSS'] = mtm_map[user_id]['max_loss']
        updated_count += 1
    elif user_id:
        not_found.append(user_id)

print(f"✅ Sheet1: {updated_count} rows updated | {len(not_found)} not found")

# Write updated Sheet1 back to workbook
for r in range(len(compiled_df)):
    for c in range(len(compiled_df.columns)):
        ws1.cell(row=r+2, column=c+1).value = compiled_df.iloc[r, c]  # +2 because header is row 1

# ====================== SHEET 2 FORMATTING ======================
print("🎨 Applying formatting to Sheet2...")

def apply_format_to_sheet2(wb):
    sheet_name = "Sheet2"
    if sheet_name not in wb.sheetnames:
        print("⚠️ Sheet2 not found!")
        return wb

    ws = wb[sheet_name]
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(bold=True, color="FFFFFF")
    grand_total_fill = PatternFill("solid", fgColor="D1C9E1")
    palette = ["F1DCDB", "ECF0DF", "F3DBDB", "DBEEF4", "FEE9D8", "B8CCE4", "FFC8CE", "F0DCDD", "ECF0E1"]
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    def norm(v):
        return str(v).strip().lower() if v is not None else ""

    def is_row_empty(ws, row_idx, max_col):
        for c in range(1, max_col + 1):
            if ws.cell(row=row_idx, column=c).value not in (None, ""):
                return False
        return True

    last_col = ws.max_column
    max_row = ws.max_row

    # Insert blank row after each Sub-Total
    sub_rows = [r for r in range(1, max_row + 1) if norm(ws.cell(r, 1).value) == "sub-total"]
    for r in reversed(sub_rows):
        after = r + 1
        if after <= max_row and is_row_empty(ws, after, last_col):
            continue
        ws.insert_rows(after, 1)

    # Detect header row
    header_row = None
    for r in range(1, min(15, ws.max_row) + 1):
        vals = [str(ws.cell(r, c).value or "").strip().upper() for c in range(1, ws.max_column + 1)]
        if "ALGO" in vals and "SERVER" in vals:
            header_row = r
            break

    if header_row is None:
        print("⚠️ Could not detect Sheet2 header")
        return wb

    # Style header
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.freeze_panes = f"A{header_row + 1}"

    # Format sections and Grand Total
    data_start = header_row + 1
    sub_rows_after = []
    grand_row = None
    for r in range(data_start, ws.max_row + 1):
        v = norm(ws.cell(r, 1).value)
        if v == "grand total":
            grand_row = r
            break
        if v == "sub-total":
            sub_rows_after.append(r)

    # Color sections
    sections = []
    start = data_start
    for s in sub_rows_after:
        sections.append((start, s))
        start = s + 2 if (s + 1 <= ws.max_row and is_row_empty(ws, s + 1, ws.max_column)) else s + 1

    for i, (rs, rend) in enumerate(sections):
        fill = PatternFill("solid", fgColor=palette[i % len(palette)])
        for rr in range(rs, rend + 1):
            for cc in range(1, ws.max_column + 1):
                cell = ws.cell(rr, cc)
                cell.fill = fill
                cell.border = thin_border

    # Bold subtotals
    for _, subtotal_r in sections:
        for c in range(1, ws.max_column + 1):
            ws.cell(subtotal_r, c).font = bold_font

    if grand_row:
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(grand_row, c)
            cell.font = bold_font
            cell.fill = grand_total_fill
            cell.border = thin_border

    # Number formatting
    return_col = None
    for c in range(1, ws.max_column + 1):
        if "RETURN" in str(ws.cell(header_row, c).value or "").upper():
            return_col = c
            break

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if isinstance(val, (int, float)):
                if c == return_col:
                    cell.value = round(float(val), 2)
                    cell.number_format = '0.00'
                else:
                    cell.value = round(val)
                    cell.number_format = '#,##0'
            cell.alignment = center_align
            cell.border = thin_border

    # Auto-fit columns
    for c in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, c).value
            if val:
                length = len(str(val))
                if ws.cell(r, c).font and ws.cell(r, c).font.bold:
                    length += 2
                max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 3, 20)

    return wb

# Apply formatting and save
wb = apply_format_to_sheet2(wb)
wb.save(output_path)

print("\n🎉 PROCESS COMPLETED SUCCESSFULLY!")
print(f"📁 Final file saved at: {output_path}")